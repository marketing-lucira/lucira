"""Multi-sheet .xlsx report (openpyxl). Summary · KPIs · Alerts · Recommendations."""
from __future__ import annotations

import io

from ..core.models import RunResult


class ExcelReporter:
    fmt = "excel"

    def render(self, result: RunResult) -> bytes:
        from openpyxl import Workbook  # lazy import
        wb = Workbook()

        ws = wb.active
        ws.title = "Summary"
        ws.append(["Lucira MIS", result.period_label])
        ws.append(["Headline", result.headline])
        ws.append([])
        ws.append(["KPI", "Value", "Prev", "Delta %", "Target", "Status"])

        kws = wb.create_sheet("KPIs")
        kws.append(["Domain", "KPI", "Dimension", "Value", "Prev", "Delta %", "Target", "Status"])
        for k in result.kpis:
            kws.append([k.metric.domain, k.metric.label, k.dim_value or "TOTAL",
                        round(k.value, 2), round(k.prev_value or 0, 2),
                        round(k.delta_pct, 1) if k.delta_pct is not None else None,
                        k.target, k.status.value])
            if k.dimension is None:
                ws.append([k.metric.label, round(k.value, 2), round(k.prev_value or 0, 2),
                           round(k.delta_pct, 1) if k.delta_pct is not None else None,
                           k.target, k.status.value])

        aws = wb.create_sheet("Alerts")
        aws.append(["Severity", "Domain", "Rule", "Message", "Owner"])
        for a in result.alerts:
            aws.append([a.severity.value, a.domain, a.rule_id, a.message, a.owner])

        rws = wb.create_sheet("Recommendations")
        rws.append(["Priority", "Owner", "Action", "ETA days", "Est. ₹ impact"])
        for r in result.recommendations:
            rws.append([r.priority, r.owner, r.action, r.eta_days, r.est_value_inr])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
